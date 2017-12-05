import sys
import os
import argparse
import tempfile
import shutil
import xml.etree.ElementTree as ET
import arcpy
from openpyxl import Workbook
from openpyxl import load_workbook
import logging

logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

AGS_HOME = arcpy.GetInstallInfo("Desktop")["InstallDir"]
METADATA_TRANSLATOR = os.path.join(AGS_HOME, r'Metadata/Translator/ARCGIS2FGDC.xml')

XLS_TAB_NAME = "dataSource"
XLS_HEADERS = []
XLS_HEADERS_FOR_UPDATE = []

NETWORK_DRIVES = []

DATA_SOURCES = {}
DATA_TARGETS = []

DATA_CATEGORIES = []
WORD_SHORTCUTS = []

DEFINED_NAMES = {}


def ascii(s):
    if s is not None:
        return s.encode('ascii', 'replace')
    else:
        return s


def get_alias_path(lDrvPath):
    for nd in NETWORK_DRIVES:
        if lDrvPath.find(nd["Path"]) == 0:
            return lDrvPath.replace(nd["Path"], nd["Drive"])
        if lDrvPath.find(nd["Drive"]) == 0:
            return lDrvPath.replace(nd["Drive"], nd["Path"])
    return lDrvPath


def load_config(configFile):
    tree = ET.parse(configFile)
    root = tree.getroot()

    for child in root.iter('xlsTab'):
        # only one tab
        XLS_TAB_NAME = child.attrib['name']

    for child in root.iter('xlsHeader'):
        XLS_HEADERS.append(child.attrib['name'])
        if 'update' in child.attrib.keys() and child.attrib['update'] == 'Y':
            XLS_HEADERS_FOR_UPDATE.append(child.attrib['name'])
    # print XLS_HEADERS
    # print XLS_HEADERS_FOR_UPDATE

    for child in root.iter('networkDrive'):
        NETWORK_DRIVES.append({
            'Drive': child.attrib['drive'],
            'Path': child.attrib['path']
        })
    # print NETWORK_DRIVES

    for child in root.iter('source'):
        nm = child.attrib['name']
        if nm not in DATA_SOURCES.keys():
            DATA_SOURCES[nm] = []
        for sc in child:
            DATA_SOURCES[nm].append({
                'LDrv':sc.attrib['path'],
                'Mode':sc.tag
            })
            DATA_SOURCES[nm].append({
                'LDrv':get_alias_path(sc.attrib['path']),
                'Mode':sc.tag
            })
    # print DATA_SOURCES

    for child in root.iter('target'):
        target = {}
        target['name'] = child.attrib["name"]
        for sc in child:
            target[sc.tag] = sc.text
        DATA_TARGETS.append(target)
    # print DATA_TARGETS

    for child in root.iter('dataCategory'):
        DATA_CATEGORIES.append({
            'Key':child.attrib['key'],
            'Name':child.attrib['name']
        })
    # print DATA_CATEGORIES

    for child in root.iter('shortcut'):
        WORD_SHORTCUTS.append({
            'Key':child.attrib['key'],
            'Name':child.attrib['name'],
            'Priority':child.attrib['priority']
        })
    # print WORD_SHORTCUTS

    for child in root.iter('definedNames'):
        scope = child.attrib['scope']
        if scope not in DEFINED_NAMES.keys():
            DEFINED_NAMES[scope] = []
        for nm in child.iter('definedName'):
            DEFINED_NAMES[scope].append({
                'Name':nm.attrib['name'],
                'Path':nm.attrib['path']
            })
            DEFINED_NAMES[scope].append({
                'Name':nm.attrib['name'],
                'Path':get_alias_path(nm.attrib['path'])
            })
    # print DEFINED_NAMES

    del root
    del tree


def get_source_type(lDrvPath):
    for k in DATA_SOURCES.keys():
        excluded = False
        ds = DATA_SOURCES[k]
        for cvt in [s for s in ds if s["Mode"] == "exclude"]:
            if lDrvPath.find(cvt["LDrv"]) == 0:
                excluded = True
                break
        if not excluded:
            for cvt in [s for s in ds if s["Mode"] == "add"]:
                if lDrvPath.find(cvt["LDrv"]) == 0:
                    return k
    return None


def get_sde_connection(target):
    for cvt in DATA_TARGETS:
        if cvt["name"] == target:
            return cvt["sdeConn"]
    return None


def get_raster_connection(target):
    for cvt in DATA_TARGETS:
        if cvt["name"] == target:
            return cvt["rasterDepot"]
    return None


def get_data_properties(lDrvPath):
    parts = lDrvPath.split("\\")
    category = parts[0]
    dataType = dataFormat = parts[1]
    if dataType not in ['gdb', 'shapefiles']:
        return dataFormat
    else:
        dataName = parts[-1]
        dataFolder = parts[-2]

        if os.path.splitext(dataName)[-1] == '.shp':
            dataType = 'shp'
        elif os.path.splitext(dataFolder)[-1] == '.gdb':
            dataType = 'gdb'


def guess_target_name(lDrvPath):
    target_source = get_source_type(lDrvPath)
    # scan the defined names
    for nm in DEFINED_NAMES[target_source]:
        if nm['Path'].lower() == lDrvPath.lower():
            return nm['Name']
    if target_source == "LNG":
        # TODO: come up with a naming convention for the LNG vendor data
        return None
    elif target_source == "MOZGIS":
        # parse the path
        for src in DATA_SOURCES[target_source]:
            if src["Mode"] == "add":
                lDrvPath = lDrvPath.replace(src["LDrv"], "")
        # parse for data type
        parts = lDrvPath.split("\\")
        category = parts[0]
        fileName = parts[-1]
        folderName = parts[-2]
        dataFormat = None
        if len(os.path.splitext(fileName)[-1]) > 0:
            dataFormat = os.path.splitext(fileName)[-1]
        elif len(os.path.splitext(folderName)[-1]) > 0:
            dataFormat = os.path.splitext(folderName)[-1]
        # determine data name
        dataName = None
        dataKeys = []
        if dataFormat == '.gdb':
            dataName = fileName
        elif dataFormat == '.shp':
            dataName = os.path.splitext(fileName)[0]
            dataName = dataName.replace(' ', '_')
        # add data category if needed
        if dataName is not None:
            for c in DATA_CATEGORIES:
                if c["Name"] == category:
                    dataKeys.append(c["Key"])
            minKey = None
            keyLen = 0
            for k in dataKeys:
                if dataName.find(k) == 0:
                    return dataName
                if minKey is None or keyLen > len(k):
                    minKey = k
                    keyLen = len(k)
            if minKey is None:
                return dataName
            else:
                return minKey + "_" + dataName

    return None


def shorten_target_name(name, maxLen=30):
    if name is None or len(name) <= maxLen:
        return name
    else:
        nameParts = name.split("_")
        shortNameParts = list(nameParts)
        shortcutPriorities = list(set([e['Priority'] for e in WORD_SHORTCUTS if e['Priority'] is not None]))
        shortcutPriorities.sort()
        shortcutPriorities.reverse()
        underLenLimit = False
        for t in shortcutPriorities:
            for p in range(0, len(nameParts)):
                nmPart = nameParts[p].upper()
                for sh in [e for e in WORD_SHORTCUTS if e['Priority'] == t]:
                    if sh["Name"].upper() == nmPart:
                        shortNameParts[p] = sh["Key"]
                        break
                underLenLimit = len("_".join(shortNameParts)) <= maxLen
                if underLenLimit:
                    break
            if underLenLimit:
                break
        # remove empty ones
        shortNameParts = [e for e in shortNameParts if len(e) > 0]
        # remove duplicate ones
        shortNameParts2 = []
        for n in range(0, len(shortNameParts)-1):
            if shortNameParts[n] != shortNameParts[n+1]:
                shortNameParts2.append(shortNameParts[n])
        shortNameParts2.append(shortNameParts[-1])
        return "_".join(shortNameParts2)


def read_from_workbook(wbPath, sheetName=None):
    wb = load_workbook(filename = wbPath, read_only=True)
    ws = wb[XLS_TAB_NAME]

    dsList = []
    # skip the first 2 rows
    r = 3
    hdrs = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

    while ws["A"+str(r)].value is not None:
        dsRecord = {}
        for c in range(0, len(XLS_HEADERS)):
            dsRecord[XLS_HEADERS[c]] = ws[hdrs[c]+str(r)].value
        dsList.append(dsRecord)
        r = r + 1

    wb.close()

    del wb

    return dsList


def update_status_in_workbook(wbPath, dsList, sheetName=None):
    wb = load_workbook(filename = wbPath)
    ws1 = wb[XLS_TAB_NAME]

    # content
    s = 2 # skip the first 2 rows
    for r in range(0, len(dsList)):
        for c in dsList[r]:
            for u in XLS_HEADERS_FOR_UPDATE:
                h = XLS_HEADERS.index(u)
                ws1.cell(row=r+s+1, column=h+1, value=dsList[r][XLS_HEADERS[h]])

    wb.save(filename = wbPath)
    wb.close()

    del wb


def update_layer_metadata(sdeFC, props, srcType):

    TEMP_DIR = tempfile.gettempdir()
    metadataFile = os.path.join(TEMP_DIR, os.path.basename(sdeFC) + '-metadata.xml')
    #migrationText = " *** Migrated from the L Drive (%s)" % props["Data Source"]
    migrationText = "<b>Retired L Drive Path: </b> %s" % props["Data Source"]
    livelinkText = ""
    if props["Livelink Link"]:
        #livelinkText = 'Click <a href="' + props["Livelink Link"] + '">here</a> to go to Livelink'
        livelinkText = '<b>Livelink Path: </b> <a href="' + props["Livelink Link"] + '">' + props["Livelink Link"] + '</a>'

    if os.path.exists(metadataFile):
        os.remove(metadataFile)

    # A- export the medata from SDE feature class
    # print 'exporting the metadata of %s to %s' % (sdeFC, metadataFile)
    arcpy.ExportMetadata_conversion(
    	Source_Metadata=sdeFC,
    	Translator=METADATA_TRANSLATOR,
    	Output_File=metadataFile
    )

    # B- modify metadata
    # print 'modifying the metadata file [%s]' % (metadataFile)
    tree = ET.parse(metadataFile)
    root = tree.getroot()
    idinfo = root.find('idinfo')
    dspt = idinfo.find('descript')
    # B1- add the element
    if dspt is None:
        dspt = ET.SubElement(idinfo, 'descript')
        ET.SubElement(dspt, 'abstract')
    else:
        abstract = dspt.find('abstract')
        if abstract is None:
            ET.SubElement(dspt, 'abstract')
    # B2- modify the element text
    abstract = dspt.find('abstract')
    if abstract.text is None:
        #abstract.text = migrationText
        abstract.text = "<![CDATA[<p/><p>%s</p>]]>" % migrationText
    elif abstract.text.find(migrationText) == -1:
        #abstract.text = "<![CDATA[%s<br/>%s<br/>%s]]>" % (abstract.text, livelinkText, migrationText)
        abstract.text = "<![CDATA[%s<p/><p>%s</p><p>%s</p>]]>" % (abstract.text, livelinkText, migrationText)

    tree.write(metadataFile)

    # C- import the modified metadata back to SDE feature class
    # print 'importing the metadata file [%s] to %s' % (metadataFile, sdeFC)
    arcpy.ImportMetadata_conversion(
    	Source_Metadata=metadataFile,
    	Import_Type="FROM_FGDC",
    	Target_Metadata=sdeFC,
    	Enable_automatic_updates="ENABLED"
    )

    # print 'The metadata of %s is updated' % sdeFC


def load_layers_in_xls(wbPath, test):
    dsList = read_from_workbook(wbPath)
    for ds in dsList:
        srcType = get_source_type(ds["Data Source"])
        if srcType is not None:
            if ds["Loaded?"] not in ["LOADED", 'EXIST']:
                if bool(ds["Verified?"]) == True:
                    if ds["Layer Type"] == "RasterLayer":
                        # find out the target workspace
                        tgt_conn = get_raster_connection(srcType)
                        if tgt_conn is not None:
                            tgt_workspace = ds["Data Source"]
                            for src in DATA_SOURCES[srcType]:
                                if src["Mode"] == "add":
                                    tgt_workspace = tgt_workspace.replace(src["LDrv"], tgt_conn)
                            ds["SDE Conn"] = os.path.dirname(tgt_workspace)
                            ds["SDE Name"] = os.path.basename(tgt_workspace)
                            if os.path.exists(tgt_workspace) or arcpy.Exists(tgt_workspace):
                                print('%-60s%s' % (ascii(ds["Name"]),"*** existing layer"))
                                ds["Loaded?"] = 'EXIST'
                            else:
                                # copy data to the target workspace
                                print('%-60s%s' % (ascii(ds["Name"]),"loading to Raster Depot as %s" % tgt_workspace))
                                if test == "test":
                                    print('%-60s%s' % (" ","^^^ TESTED"))
                                    ds["Loaded?"] = 'TESTED'
                                else:
                                    if ds["SDE Conn"].lower().endswith('.gdb') == True:
                                        # copy the whole GDB
                                        if os.path.exists(ds["SDE Conn"]):
                                            shutil.rmtree(ds["SDE Conn"])
                                        # arcpy.Copy_management(os.path.dirname(ds["Data Source"]), ds["SDE Conn"])
                                        shutil.copytree(os.path.dirname(ds["Data Source"]), ds["SDE Conn"])
                                    else:
                                        try:
                                            if not os.path.exists(ds["SDE Conn"]):
                                                os.makedirs(ds["SDE Conn"])
                                            # copy the raster file set
                                            arcpy.CopyRaster_management(in_raster=ds["Data Source"], out_rasterdataset=tgt_workspace)
                                            print('%-60s%s' % (" ","^^^ LOADED"))
                                            ds["Loaded?"] = 'LOADED'
                                        except:
                                            print('%-60s%s' % (" ",">>> FAILED to load data [" + str(arcpy.GetMessages()) + "]"))
                                            ds["Loaded?"] = 'FAILED'
                        else:
                            print('%-60s%s' % (ascii(ds["Name"]),"*** no target store"))
                            ds["Loaded?"] = "NO TARGET STORE"

                    elif ds["Layer Type"] is not None and ds["Layer Type"] == "FeatureLayer":
                        tgt_conn = get_sde_connection(srcType)
                        if tgt_conn is not None:
                            ds["SDE Conn"] = tgt_conn
                            if ds["SDE Name"] is None or len(ds["SDE Name"]) > 30:
                                ds["SDE Name"] = guess_target_name(ds["Data Source"])
                            if ds["SDE Name"] is not None and len(ds["SDE Name"]) > 0:
                                ds["SDE Name"] = shorten_target_name(ds["SDE Name"])
                                if len(ds["SDE Name"]) > 30:
                                    print('%-60s%s' % (ascii(ds["Name"]),"*** name too long [%s]" % ds["SDE Name"]))
                                    ds["Loaded?"] = "NAME TOO LONG"
                                elif arcpy.Exists(tgt_conn + "\\" + ds["SDE Name"]) == True:
                                    print('%-60s%s' % (ascii(ds["Name"]),"*** existing layer"))
                                    ds["Loaded?"] = 'EXIST'
                                else:
                                    print('%-60s%s' % (ascii(ds["Name"]),"loading to SDE at %s as %s" % (tgt_conn, ds["SDE Name"])))
                                    # upload the actual data
                                    if test == "test":
                                        print('%-60s%s' % (" ","^^^ TESTED"))
                                        ds["Loaded?"] = 'TESTED'
                                    else:
                                        try:
                                            arcpy.CopyFeatures_management(ds["Data Source"], tgt_conn + "\\" + ds["SDE Name"])
                                            print('%-60s%s' % (" ","^^^ LOADED"))
                                            ds["Loaded?"] = 'LOADED'

                                            print('%-60s%s' % (ascii(ds["Name"]),"updating SDE metadata of %s\\%s" % (tgt_conn, ds["SDE Name"])))
                                            update_layer_metadata(tgt_conn + "\\" + ds["SDE Name"], ds, srcType)
                                            print('%-60s%s' % (" ","^^^ METADATA UPDATED"))
                                        except arcpy.ExecuteError:
                                            print('%-60s%s' % (" ",">>> FAILED to load data [" + str(arcpy.GetMessages()) + "]"))
                                            ds["Loaded?"] = 'FAILED'
                            else:
                                print('%-60s%s' % (ascii(ds["Name"]),"*** no target name [" + ds["Data Source"] + "]"))
                                ds["Loaded?"] = "NO TARGET NAME"
                        else:
                            print('%-60s%s' % (ascii(ds["Name"]),"*** no target store  [" + ds["Data Source"] + "]"))
                            ds["Loaded?"] = "NO TARGET STORE"
                    else:
                        print('%-60s%s' % (ascii(ds["Name"]),"*** unknown layer type"))
                        ds["Loaded?"] = "UNKNOWN LAYER"
                else:
                    print('%-60s%s' % (ascii(ds["Name"]),"*** invalid layer"))
                    ds["Loaded?"] = "INVALID"
            else:
                print('%-60s%s' % (ascii(ds["Name"]),"*** existing layer"))
                # ds["Loaded?"] = 'EXIST'
        else:
            print('%-60s%s' % (ascii(ds["Name"]),"*** out of scope"))
            ds["Loaded?"] = "NOT IN SCOPE"

    return dsList


def load_layers_in_file(xlsPath, test):
    print('\nThe xlsx file: %s' % xlsPath)
    dsList = load_layers_in_xls(xlsPath, test)
    # update status in the xls file
    update_status_in_workbook(xlsPath, dsList)


def load_layers_in_folder(xlsFolder, test):
    for root, dirs, files in os.walk(xlsFolder):
        # walk through all files
        for fname in files:
            if fname.endswith(".xlsx") and not fname.startswith('~$'):
                # read from a xls file
                xlsPath = os.path.join(root, fname)
                load_layers_in_file(xlsPath, test)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Import data referenced by layers in spreadsheets')
    parser.add_argument('-x','--xls', help='XLS Folder/File (input)', required=True)
    parser.add_argument('-a','--action', help='Action Options (batch, single, test)', required=False, default='batch')
    parser.add_argument('-c','--cfg', help='Config File', required=False, default=r'H:\MXD_Scan\config.xml')

    params = parser.parse_args()

    if params.cfg is not None:
        load_config(params.cfg)

    if params.action == 'batch':
        load_layers_in_folder(params.xls, None)
    elif params.action == 'single':
        load_layers_in_file(params.xls, None)
    elif params.action == 'test':
        load_layers_in_folder(params.xls, 'test')
    else:
        print 'Error: unknown action [%s] for importing' % params.action

