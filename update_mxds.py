import sys
import os
import argparse
import re
import shutil
#import xml.etree.ElementTree as ET
from lxml import etree as ET
from arcpy import mapping
from openpyxl import Workbook
from openpyxl import load_workbook
import logging

logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

XLS_TAB_NAME = "dataSource"
XLS_HEADERS = []


def ascii(s):
    if s is not None:
        return s.encode('ascii', 'replace')
    else:
        return s


def load_config(configFile):
    tree = ET.parse(configFile)
    root = tree.getroot()

    for child in root.iter('xlsTab'):
        # only one tab
        XLS_TAB_NAME = child.attrib['name']

    for child in root.iter('xlsHeader'):
        XLS_HEADERS.append(child.attrib['name'])

    del root
    del tree


def read_from_workbook(wbPath, sheetName=None):
    wb = load_workbook(filename = wbPath, read_only=True)
    ws = wb[XLS_TAB_NAME]

    dsList = []
    # skip the first 2 rows
    r = 3
    hdrs = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

    while ws["A"+str(r)].value is not None:
        dsRecord = {}
        for c in range(0, len(XLS_HEADERS)):
            dsRecord[XLS_HEADERS[c]] = ws[hdrs[c]+str(r)].value
        dsList.append(dsRecord)
        r = r + 1

    wb.close()

    del wb

    return dsList


def update_mxd_ds(mxdPath, xlsPath, newMxdPath):
    print('\nThe mxd file: %s' % mxdPath)
    # create the target folder if not exist yet
    newMxdFolder = os.path.dirname(newMxdPath)
    if not os.path.exists(newMxdFolder):
        os.makedirs(newMxdFolder)
    # process the mxd file
    mxd = None
    lyrUpdateCount = 0
    try:
        mxd = arcpy.mapping.MapDocument(mxdPath)
        lyrs = arcpy.mapping.ListLayers(mxd)

        dsList = read_from_workbook(xlsPath)

        print('Updating data sources ...')
        for lyr in lyrs:
             if lyr.isFeatureLayer == True:
                 lyrType = "FeatureLayer"
             elif lyr.isRasterLayer == True:
                 lyrType = "RasterLayer"
             elif lyr.isServiceLayer == True:
                 lyrType = "ServiceLayer"
             elif lyr.isGroupLayer == True:
                 lyrType = "GroupLayer"
             else:
                 lyrType = "UnknownLayer"
             try:
                 if lyrType in ["GroupLayer", "ServiceLayer", "UnknownLayer"]:
                     print('%-60s%s' % (ascii(lyr.name),"*** skip " + lyrType))
                 else:
                     # get the layer data source
                     lyrDS = lyr.dataSource
                     for ds in dsList:
                        if ds["Data Source"] == lyrDS:
                            if bool(ds["Verified?"]) == True and ds["Loaded?"] in ["LOADED", 'EXIST']:
                                if ds["Layer Type"] == "RasterLayer":
                                    wsType = "RASTER_WORKSPACE"
                                elif ds["Layer Type"] == "FeatureLayer":
                                    wsType = "SDE_WORKSPACE"
                                else:
                                    wsType = "NONE"
                                try:
                                    # update the layer data source
                                    print('%-60s%s' % (ascii(lyr.name),"set data source to " + os.path.join(ds["SDE Conn"], ds["SDE Name"])))
                                    lyr.replaceDataSource(ds["SDE Conn"], wsType, ds["SDE Name"], True)
                                    lyrUpdateCount = lyrUpdateCount + 1
                                except:
                                    print('%-60s%s' % (ascii(lyr.name),">>> failed to update data source " + lyrType))
                            else:
                                print('%-60s%s' % (ascii(lyr.name),"*** skip invalid or unavailable layer"))
                            break

             except:
                 print('%-60s%s' % (ascii(lyr.name),">>> failed to retrieve info " + lyrType))
        if lyrUpdateCount > 0:
            # save the updated mxd file
            mxd.saveACopy(newMxdPath)
            print('\nThe NEW mxd file: %s' % newMxdPath)
        else:
            print('No layer updated in the mxd file [%s]' % mxdPath)
            shutil.copy2(mxdPath, newMxdPath)
            print('\nThe COPIED mxd file copied as is: %s' % newMxdPath)

    except:
        print('Unable to open the mxd file [%s]' % mxdPath)
        shutil.copy2(mxdPath, newMxdPath)
        print('\nThe COPIED mxd file copied as is: %s' % newMxdPath)

    finally:
        del mxd


def update_mxds(mxd_folder, xls_folder, new_mxd_folder, date_filters):
    # parse date filters
    lower_date = None
    upper_date = None
    if date_filters is not None:
        dfs = re.split('<', date_filters)
        if len(dfs) == 2:
            if dfs[0] is not None and len(dfs[0]) > 0:
                lower_date = datetime.datetime.strptime(dfs[0], "%Y/%m/%d")
            if dfs[1] is not None and len(dfs[1]) > 0:
                upper_date = datetime.datetime.strptime(dfs[1], "%Y/%m/%d")

    # walk through the mxd files
    for mxdRoot, mxdDirs, mxdFiles in os.walk(mxd_folder):
        for mxdFile in mxdFiles:
            if mxdFile.endswith(".mxd"):
                mxdPath = os.path.join(mxdRoot, mxdFile)

                modified_time = datetime.datetime.fromtimestamp(os.stat(mxdPath).st_mtime)
                # check against the filter
                is_filter_met = True
                if lower_date is not None:
                    is_filter_met = modified_time >= lower_date
                if upper_date is not None:
                    is_filter_met = is_filter_met and modified_time <= upper_date

                # work on the mxd file
                if is_filter_met == True:
                    newMxdPath = mxdPath.replace(mxd_folder, new_mxd_folder)
                    #newMxdPath = os.path.join(new_mxd_folder, mxdPath.replace(mxd_folder, new_mxd_folder))

                    xlsPath = os.path.join(xls_folder, mxdPath.replace(mxd_folder, xls_folder))
                    xlsPath = xlsPath + ".xlsx"

                    if not os.path.exists(xlsPath):
                        print('Failed to process the mxd file [%s] without the xls file [%s]' % (mxdPath, xlsPath))
                    else:
                        update_mxd_ds(mxdPath, xlsPath, newMxdPath)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Update mxds with data sources specified in spreadsheets')
    parser.add_argument('-m','--mxd', help='MXD Folder/File (input)', required=True)
    parser.add_argument('-x','--xls', help='XLS Folder/File (input)', required=True)
    parser.add_argument('-n','--newMxd', help='New MXD Folder/File (output)', required=True)
    parser.add_argument('-f','--filter', help='Filter by dates. Ex. 2015/2/19<2016/2/18', required=False, default=None)
    parser.add_argument('-a','--action', help='Action Options (batch, single)', required=False, default='batch')
    parser.add_argument('-c','--cfg', help='Config File', required=False, default=r'H:\MXD_Scan\config.xml')

    params = parser.parse_args()

    if params.cfg is not None:
        load_config(params.cfg)

    if params.action == 'batch':
        update_mxds(params.mxd, params.xls, params.newMxd, params.filter)
        print('###### Completed updating all mxd files in %s'%params.mxd)
    elif params.action == 'single':
        update_mxd_ds(params.mxd, params.xls, params.newMxd)
        print('###### Completed updating the mxd file at %s'%params.mxd)
    else:
        print 'Error: unknown action [%s] for scanning' % params.action

