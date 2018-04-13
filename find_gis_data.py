import sys
import os
import argparse
import tempfile
import shutil
import xml.etree.ElementTree as ET
import arcpy
from openpyxl import Workbook
from openpyxl import load_workbook
import logging


logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

AGS_HOME = arcpy.GetInstallInfo("Desktop")["InstallDir"]
METADATA_TRANSLATOR = os.path.join(AGS_HOME, r'Metadata/Translator/ARCGIS2FGDC.xml')

XLS_TAB_NAME = "dataSource"
XLS_HEADERS = []

NETWORK_DRIVES = []
SNAPSHOT_PATHS = []

DATA_SOURCES = {}

XREF_TAB_NAME = "xRef"
XREF_HEADERS = ["Source Type", "Data Source", "SDE Path"]

OUTPUT_TAB_NAME = "xRef"
OUTPUT_HEADERS = XREF_HEADERS + ["Loaded In SDE?", "Snapshot Path"]


def ascii(s):
    if s is not None:
        return s.encode('ascii', 'replace')
    else:
        return s


def get_alias_path(lDrvPath):
    for nd in NETWORK_DRIVES:
        if lDrvPath.find(nd["Path"]) == 0:
            return lDrvPath.replace(nd["Path"], nd["Drive"])
        if lDrvPath.find(nd["Drive"]) == 0:
            return lDrvPath.replace(nd["Drive"], nd["Path"])
    return lDrvPath


def path_unc_2_md(lDrvPath):
    for nd in NETWORK_DRIVES:
        if lDrvPath.find(nd["Path"]) == 0:
            return lDrvPath.replace(nd["Path"], nd["Drive"])
    if re.match("^\\\\[a-zA-z]{1}", lDrvPath[0:3]) is not None:
        return lDrvPath
    else:
        return None


def path_md_2_unc(lDrvPath):
    for nd in NETWORK_DRIVES:
        if lDrvPath.find(nd["Drive"]) == 0:
            return lDrvPath.replace(nd["Drive"], nd["Path"])
    if re.match("^[a-zA-z]{1}\:\\\\", lDrvPath[0:3]) is not None:
        return lDrvPath
    return lDrvPath


def path_snapshot_2_source(lDrvPath):
    for sp in SNAPSHOT_PATHS:
        if lDrvPath.find(sp["Snapshot"]) == 0:
            return lDrvPath.replace(sp["Snapshot"], sp["Path"])
    return lDrvPath


def load_config(configFile):
    tree = ET.parse(configFile)
    root = tree.getroot()

    for child in root.iter('xlsTab'):
        # only one tab
        XLS_TAB_NAME = child.attrib['name']

    for child in root.iter('xlsHeader'):
        XLS_HEADERS.append(child.attrib['name'])
    # print XLS_HEADERS

    for child in root.iter('networkDrive'):
        NETWORK_DRIVES.append({
            'Drive': child.attrib['drive'],
            'Path': child.attrib['path']
        })
    # print NETWORK_DRIVES

    for child in root.iter('snapshotPath'):
        SNAPSHOT_PATHS.append({
            'Path': child.attrib['path'],
            'Snapshot': child.attrib['snapshot']
        })
    # print SNAPSHOT_PATHS

    for child in root.iter('source'):
        nm = child.attrib['name']
        if nm not in DATA_SOURCES.keys():
            DATA_SOURCES[nm] = []
        for sc in child:
            DATA_SOURCES[nm].append({
                'LDrv':sc.attrib['path'],
                'Mode':sc.tag
            })
##            DATA_SOURCES[nm].append({
##                'LDrv':get_alias_path(sc.attrib['path']),
##                'Mode':sc.tag
##            })
    # print DATA_SOURCES


def read_from_workbook(wbPath, sheetName=None):
    wb = load_workbook(filename = wbPath, read_only=True)
    ws = wb[XREF_TAB_NAME]

    dsList = []
    # skip the first row
    r = 2
    hdrs = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

    while ws["A"+str(r)].value is not None:
        dsRecord = {}
        for c in range(0, len(XREF_HEADERS)):
            dsRecord[XREF_HEADERS[c]] = ws[hdrs[c]+str(r)].value
        dsList.append(dsRecord)
        r = r + 1

    wb.close()

    del wb

    return dsList


def write_to_workbook(wbPath, dsList, sheetName=None):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = OUTPUT_TAB_NAME

    # headers
    for c in range(0, len(OUTPUT_HEADERS)):
        ws1.cell(row=1, column=c+1, value=OUTPUT_HEADERS[c])

    # content
    s = 1 # skip the first 1 row
    for r in range(0, len(dsList)):
        for c in dsList[r]:
            # TODO: add style to cell?
            h = OUTPUT_HEADERS.index(c)
            if h > -1:
                ws1.cell(row=r+s+1, column=h+1, value=dsList[r][OUTPUT_HEADERS[h]])
            else:
                print('Invalid header [%s] in xls [%s]' % (c, mxdPath))

    wb.save(filename = wbPath)
    wb.close()

    del wb


def get_source_type(lDrvPath):
    for k in DATA_SOURCES.keys():
        excluded = False
        ds = DATA_SOURCES[k]
        for cvt in [s for s in ds if s["Mode"] == "exclude"]:
            if lDrvPath.find(cvt["LDrv"]) == 0:
                excluded = True
                break
        if not excluded:
            for cvt in [s for s in ds if s["Mode"] == "add"]:
                if lDrvPath.find(cvt["LDrv"]) == 0:
                    return k
    return None


def get_sde_path(filePath, sdeDataList):
    for sd in sdeDataList:
        if filePath == sd["Data Source"] or filePath == get_alias_path(sd["Data Source"]):
            return sd["SDE Path"]
    return None


def onWalkError(osError):
    print('#### Error on %s' % str(osError))


def scan_data_in_LDrive(dataFolder, fileExts):
    dataFileList = []
    srcType = get_source_type(dataFolder)
    if srcType is not None:
        try:
            #arcWalk = arcpy.da.Walk(dataFolder, onerror=None, datatype=fileExts)
            arcWalk = arcpy.da.Walk(dataFolder, onerror=onWalkError, datatype=fileExts)
            for root, dirs, files in arcWalk:
            #for root, dirs, files in os.walk(dataFolder):
                if root.endswith(".gdb"):
                    continue    # don't go into a gdb folder
                for fname in files:
                    fbase,fext = os.path.splitext(fname)
                    if fext in [".zip", ".doc", ".docx", ".pdf", ".mxd"]:
                        continue    # skip non-related files
                    snapshotPath = os.path.join(root, fname)
                    print("Found: %s" % ascii(snapshotPath))
                    sourcePath = path_snapshot_2_source(snapshotPath)
                    sourceUncPath = path_md_2_unc(sourcePath)
                    snapshotUncPath = path_md_2_unc(snapshotPath)
                    dataFileList.append({
                        "Source Type": srcType,
                        "Data Source": sourceUncPath,
                        "Snapshot Path": snapshotUncPath
                    })
        except:
            print "#### Unexpected error:", sys.exc_info()[0]

    return dataFileList


def compile_data_with_SDE(dataFolder, fileExts, xlsOutput, xlsInput):

    dataFileList = scan_data_in_LDrive(dataFolder, fileExts)
    if len(dataFileList) == 0:
        print('#### No GIS data found: %s' % dataFolder)
    else:
        sdeDataList = read_from_workbook(xlsInput)

        for df in dataFileList:
            sdePath = get_sde_path(df["Data Source"], sdeDataList)
            df["SDE Path"] = sdePath
            df["Loaded In SDE?"] = ('No' if sdePath is None else 'Yes')

    outputDir = os.path.dirname(xlsOutput)
    if not os.path.exists(outputDir):
        os.makedirs(outputDir)
    write_to_workbook(xlsOutput, sorted(dataFileList, key=lambda df:df["Source Type"]))

    print('**** The compiled data list: %s' % xlsOutput)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Find all ESRI-compatible data')
    parser.add_argument('-d','--dir', help='Data Directory to Scan (input)', required=True)
    parser.add_argument('-o','--dsf', help='XLS File (output)', required=True)
    # Esri_default="FeatureClass,Layer,CadDrawing,RasterCatalog,RasterDataset"
    parser.add_argument('-t','--typ', help='Data Type (input)', required=False, default="FeatureClass,RasterCatalog,RasterDataset")
    # fext_default='.shp,.gdb,.dwg,.dgn'
    #parser.add_argument('-t','--typ', help='Data Type (input)', required=False, default=".shp,.gdb")
    parser.add_argument('-x','--xls', help='XLS File (input)', required=False, default=r"C:\Users\kdb086\Documents\scan_all\LDrive2SDE_Data_Path.xlsx")
    parser.add_argument('-c','--cfg', help='Config File', required=False, default=r'C:\Users\kdb086\Documents\scan_all\config_snp.xml')

    params = parser.parse_args()

    if params.cfg is not None:
        load_config(params.cfg)

    gisDataTypes = []
    if params.typ is not None:
        gisDataTypes = params.typ.split(',')

    compile_data_with_SDE(params.dir, gisDataTypes, params.dsf, params.xls)
