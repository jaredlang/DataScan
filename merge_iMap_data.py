import sys
import os
import argparse
import tempfile
import shutil
import xml.etree.ElementTree as ET
import arcpy
from openpyxl import Workbook
from openpyxl import load_workbook
import logging

logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

XLS_TAB_NAME = "dataSource"
XLS_HEADERS = []
XLS_HEADERS_FOR_UPDATE = []

NETWORK_DRIVES = []

DATA_SOURCES = {}
DATA_TARGETS = []

DATA_CATEGORIES = []
WORD_SHORTCUTS = []

DEFINED_NAMES = {}

MAPPER_TAB_NAME = "GIS"
MAPPER_HEADERS = ['Name_in_mxd', 'Source_Featureclass_Name', 'Source_Path', 'Target_Path', 'Target_Featureclass_Name']
MAPPER_LIST = []

MAPPED_HEADERS_FOR_UPDATE = ["Data Source", "Verified?", "SDE Name", "SDE Conn"]


def get_alias_path(lDrvPath):
    for nd in NETWORK_DRIVES:
        if lDrvPath.find(nd["Path"]) == 0:
            return lDrvPath.replace(nd["Path"], nd["Drive"])
        if lDrvPath.find(nd["Drive"]) == 0:
            return lDrvPath.replace(nd["Drive"], nd["Path"])
    return lDrvPath


def load_config(configFile):
    tree = ET.parse(configFile)
    root = tree.getroot()

    for child in root.iter('xlsTab'):
        # only one tab
        XLS_TAB_NAME = child.attrib['name']

    for child in root.iter('xlsHeader'):
        XLS_HEADERS.append(child.attrib['name'])
        if 'update' in child.attrib.keys() and child.attrib['update'] == 'Y':
            XLS_HEADERS_FOR_UPDATE.append(child.attrib['name'])
    # print XLS_HEADERS
    # print XLS_HEADERS_FOR_UPDATE

    for child in root.iter('networkDrive'):
        NETWORK_DRIVES.append({
            'Drive': child.attrib['drive'],
            'Path': child.attrib['path']
        })
    # print NETWORK_DRIVES

    for child in root.iter('source'):
        nm = child.attrib['name']
        if nm not in DATA_SOURCES.keys():
            DATA_SOURCES[nm] = []
        for sc in child:
            DATA_SOURCES[nm].append({
                'LDrv':sc.attrib['path'],
                'Mode':sc.tag
            })
            DATA_SOURCES[nm].append({
                'LDrv':get_alias_path(sc.attrib['path']),
                'Mode':sc.tag
            })
    # print DATA_SOURCES

    for child in root.iter('target'):
        target = {}
        target['name'] = child.attrib["name"]
        for sc in child:
            target[sc.tag] = sc.text
        DATA_TARGETS.append(target)
    # print DATA_TARGETS

    for child in root.iter('dataCategory'):
        DATA_CATEGORIES.append({
            'Key':child.attrib['key'],
            'Name':child.attrib['name']
        })
    # print DATA_CATEGORIES

    for child in root.iter('shortcut'):
        WORD_SHORTCUTS.append({
            'Key':child.attrib['key'],
            'Name':child.attrib['name'],
            'Priority':child.attrib['priority']
        })
    # print WORD_SHORTCUTS

    for child in root.iter('definedNames'):
        scope = child.attrib['scope']
        if scope not in DEFINED_NAMES.keys():
            DEFINED_NAMES[scope] = []
        for nm in child.iter('definedName'):
            DEFINED_NAMES[scope].append({
                'Name':nm.attrib['name'],
                'Path':nm.attrib['path']
            })
            DEFINED_NAMES[scope].append({
                'Name':nm.attrib['name'],
                'Path':get_alias_path(nm.attrib['path'])
            })
    # print DEFINED_NAMES

    del root
    del tree


def get_source_type(lDrvPath):
    for k in DATA_SOURCES.keys():
        excluded = False
        ds = DATA_SOURCES[k]
        for cvt in [s for s in ds if s["Mode"] == "exclude"]:
            if lDrvPath.find(cvt["LDrv"]) == 0:
                excluded = True
                break
        if not excluded:
            for cvt in [s for s in ds if s["Mode"] == "add"]:
                if lDrvPath.find(cvt["LDrv"]) == 0:
                    return k
    return None


def get_sde_connection(target):
    for cvt in DATA_TARGETS:
        if cvt["name"] == target:
            return cvt["sdeConn"]
    return None


def get_raster_connection(target):
    for cvt in DATA_TARGETS:
        if cvt["name"] == target:
            return cvt["rasterDepot"]
    return None


def get_data_properties(lDrvPath):
    parts = lDrvPath.split("\\")
    category = parts[0]
    dataType = dataFormat = parts[1]
    if dataType not in ['gdb', 'shapefiles']:
        return dataFormat
    else:
        dataName = parts[-1]
        dataFolder = parts[-2]

        if os.path.splitext(dataName)[-1] == '.shp':
            dataType = 'shp'
        elif os.path.splitext(dataFolder)[-1] == '.gdb':
            dataType = 'gdb'


def verify_layer_dataSource(lyr):
    try:
        return arcpy.Exists(lyr)
    except:
        logger.error("failed to open layer [%s]" % fc)
        return False


def read_from_workbook(wbPath, xlsHeaders, sheetName, skipRows):
    wb = load_workbook(filename = wbPath, read_only=True)
    ws = wb[sheetName]

    dsList = []
    # skip the first 2 rows
    r = skipRows + 1
    hdrs = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

    while ws["A"+str(r)].value is not None:
        dsRecord = {}
        for c in range(0, len(xlsHeaders)):
            dsRecord[xlsHeaders[c]] = ws[hdrs[c]+str(r)].value
        dsList.append(dsRecord)
        r = r + 1

    wb.close()

    del wb

    return dsList


def update_status_in_workbook(wbPath, dsList, xlsHeaders, xlsHeadersForUpdate, sheetName, skipRows):
    wb = load_workbook(filename = wbPath)
    ws1 = wb[sheetName]

    # content
    s = skipRows # skip the first 2 rows
    for r in range(0, len(dsList)):
        for c in dsList[r]:
            for u in xlsHeadersForUpdate:
                h = xlsHeaders.index(u)
                ws1.cell(row=r+s+1, column=h+1, value=dsList[r][xlsHeaders[h]])

    wb.save(filename = wbPath)
    wb.close()

    del wb


def parse_data_name(dataSource):
    dsParts = dataSource.split('.')
    return dsParts[-1]


def get_predefined_name(dataName):
    for mp in MAPPER_LIST:
        if mp["Target_Featureclass_Name"] == dataName:
            return mp["Target_Featureclass_Name"]
    return None

def get_llDrv_path(dataName):
    for mp in MAPPER_LIST:
        if mp["Target_Featureclass_Name"] == dataName:
            return os.path.join(mp["Source_Path"], mp["Source_Featureclass_Name"])
    return None


def map_layers_in_xls(wbPath):
    dsList = read_from_workbook(wbPath, XLS_HEADERS, XLS_TAB_NAME, 2)
    for ds in dsList:
        if bool(ds["Verified?"]) == False:
            if ds["Layer Type"] == "FeatureLayer":
                dataName = parse_data_name(ds["Data Source"])
                llDrvPath = get_llDrv_path(dataName)
                if llDrvPath is None:
                    print('%-60s%s' % (ds["Name"],"No LDrive path found for %s" % ds["Data Source"]))
                else:
                    ds["Data Source"] = llDrvPath
                    ds["Verified?"] = verify_layer_dataSource(ds["Data Source"])
                    '''
                    srcType = get_source_type(ds["Data Source"])
                    if srcType == "MOZGIS":
                        ds["SDE Name"] = get_predefined_name(dataName)
                        ds["SDE Conn"] = get_sde_connection(srcType)
                    '''

    return dsList


def map_layers_in_file(xlsPath):
    print('\nThe xlsx file: %s' % xlsPath)
    dsList = map_layers_in_xls(xlsPath)
    # update status in the xls file
    update_status_in_workbook(xlsPath, dsList, XLS_HEADERS, MAPPED_HEADERS_FOR_UPDATE, XLS_TAB_NAME, 2)


def map_files_in_folder(xlsFolder, mapperPath):
    MAPPER_LIST.extend(read_from_workbook(mapperPath, MAPPER_HEADERS, MAPPER_TAB_NAME, 1))
    for root, dirs, files in os.walk(xlsFolder):
        # walk through all files
        for fname in files:
            if fname.endswith(".xlsx") and not fname.startswith('~$'):
                # read from a xls file
                xlsPath = os.path.join(root, fname)
                map_layers_in_file(xlsPath)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Merge iMap data into MOZGIS and LNG')
    parser.add_argument('-x','--xls', help='XLS Folder (input)', required=True)
    parser.add_argument('-m','--map', help='Data Path Mapper (input)', required=True)
    parser.add_argument('-c','--cfg', help='Config File', required=False, default=r'H:\MXD_Scan\config.xml')

    params = parser.parse_args()

    if params.cfg is not None:
        load_config(params.cfg)

    map_files_in_folder(params.xls, params.map)
