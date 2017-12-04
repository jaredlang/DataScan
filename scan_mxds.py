import sys
import os
import argparse
import re
import xml.etree.ElementTree as ET
from datetime import datetime
import arcpy
from openpyxl import Workbook
from openpyxl import load_workbook
import logging

logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

XLS_TAB_NAME = "dataSource"
XLS_HEADERS = []

NETWORK_DRIVES = []

DATA_SOURCES = []
DATA_TARGETS = []

XREF_TAB_NAME = "xRef"
XREF_HEADERS = []
LDRV_LL_PATHS = []


def ascii(s):
    if s is not None:
        return s.encode('ascii', 'replace')
    else:
        return s


def get_alias_path(lDrvPath):
    for nd in NETWORK_DRIVES:
        if lDrvPath.find(nd["Path"]) == 0:
            return lDrvPath.replace(nd["Path"], nd["Drive"])
        if lDrvPath.find(nd["Drive"]) == 0:
            return lDrvPath.replace(nd["Drive"], nd["Path"])
    return lDrvPath


def load_config(configFile):
    tree = ET.parse(configFile)
    root = tree.getroot()

    for child in root.iter('xlsTab'):
        # only one tab
        XLS_TAB_NAME = child.attrib['name']

    for child in root.iter('xlsHeader'):
        XLS_HEADERS.append(child.attrib['name'])
    # print XLS_HEADERS

    for child in root.iter('source'):
        src_name = child.attrib['name']
        for sc in child:
            DATA_SOURCES.append({
                'name': src_name,
                'LDrv':sc.attrib['path'],
                'Mode':sc.tag
            })
            DATA_SOURCES.append({
                'name': src_name,
                'LDrv':get_alias_path(sc.attrib['path']),
                'Mode':sc.tag
            })
    # print DATA_SOURCES

    for child in root.iter('target'):
        target = {}
        target['name'] = child.attrib['name']
        for sc in child:
            target[sc.tag] = sc.text
            DATA_TARGETS.append(target)
    # print DATA_TARGETS

    '''
    for src in DATA_SOURCES:
        for tgt in DATA_TARGETS:
            if src['name'] == tgt['name'] and src['Mode'] == 'add':
                LDRV_LL_PATHS.append({
                    'Name':src['name'],
                    'LDrv':src['LDrv'],
                    'LL':tgt['livelink']
                })
    '''
    for child in root.iter('livelink'):
        if child.attrib['name'] == 'xRef':
            llPath = child.attrib['path']
            for hdr in child.iter('xrefHeader'):
                XREF_HEADERS.append(hdr.attrib['name'])

            wb = load_workbook(filename = llPath, read_only=True)
            ws = wb[XREF_TAB_NAME]

            # skip the first 1 row (the headers)
            r = 2
            hdrs = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

            while ws["A"+str(r)].value is not None:
                xrefRecord = {}
                for c in range(0, len(XREF_HEADERS)):
                    xrefRecord[XREF_HEADERS[c]] = ws[hdrs[c]+str(r)].value
                LDRV_LL_PATHS.append(xrefRecord)
                r = r + 1

            wb.close()
            del wb

            break

    # print LDRV_LL_PATHS


def find_Livelink_path(lDrvPath):
    for cvt in LDRV_LL_PATHS:
        if lDrvPath.find(cvt["Data Source"]) == 0:
            return cvt["Livelink URL"]
    return None


def verify_layer_dataSource(lyr, lyrType):
    try:
        return arcpy.Exists(lyr)
    except:
        logger.error("failed to open layer [%s]" % fc)
        return False


def write_to_workbook(wbPath, dsList, sheetName=None):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = XLS_TAB_NAME

    # headers
    for c in range(0, len(XLS_HEADERS)):
        ws1.cell(row=1, column=c+1, value=XLS_HEADERS[c])

    # content
    s = 1 # skip the first 1 row
    for r in range(0, len(dsList)):
        for c in dsList[r]:
            # TODO: add style to cell?
            h = XLS_HEADERS.index(c)
            if h > -1:
                ws1.cell(row=r+s+1, column=h+1, value=dsList[r][XLS_HEADERS[h]])
            else:
                print('Invalid header [%s] in xls [%s]' % (c, mxdPath))

    wb.save(filename = wbPath)
    wb.close()

    del wb


def scan_layers_in_mxd(mxdPath):
     dsList = []
     mxd = None
     # scan the mxd file
     try:
         mxd = arcpy.mapping.MapDocument(mxdPath)
         lyrs = arcpy.mapping.ListLayers(mxd)
         dsList.append({"Name": "MXD File", "Data Source": mxdPath})
         lyrTitle = 'Map Layer'
         sourceTitle = 'Data Source'
         print('%-60s%s' % (lyrTitle,sourceTitle))
         print('%-60s%s' % ('-' * len(lyrTitle), '-' * len(sourceTitle)))
         for lyr in lyrs:
             if lyr.isFeatureLayer == True:
                 lyrType = "FeatureLayer"
             elif lyr.isRasterLayer == True:
                 lyrType = "RasterLayer"
             elif lyr.isServiceLayer == True:
                 lyrType = "ServiceLayer"
             elif lyr.isGroupLayer == True:
                 lyrType = "GroupLayer"
             else:
                 lyrType = "UnknownLayer"
             try:
                 if lyrType in ["GroupLayer", "ServiceLayer", "UnknownLayer"]:
                     print('%-60s%s' % (ascii(lyr.name),"*** skip " + lyrType))
                 else:
                     # get the layer data source
                     ds = lyr.dataSource
                     print('%-60s%s' % (ascii(lyr.name),ds))
                     # get the layer def query
                     defQry = None
                     if lyr.supports("DEFINITIONQUERY") == True:
                         defQry = lyr.definitionQuery
                     # get the layer description
                     dspt = None
                     if lyr.supports("DESCRIPTION") == True:
                         dspt = lyr.description
                     #
                     # verify the layer data source
                     verified = verify_layer_dataSource(ds, lyrType)
                     # get the Livelink path
                     llPath = find_Livelink_path(ds)
                     if llPath is None:
                        print('%-60s%s' % (ascii(lyr.name),"??? no Livelink path found for " + ds))
                     #
                     dsList.append({"Name": lyr.name, "Data Source": ds, "Layer Type": lyrType,
                                    "Verified?": verified, "Definition Query": defQry, "Description": dspt,
                                    "Livelink Link": llPath})
             except:
                 print('%-60s%s' % (ascii(lyr.name), ">>> failed to retrieve info " + lyrType + ": " + str(sys.exc_info()[0])))
     except:
         print('Failed to process the mxd file [%s]: %s' % (mxdPath, sys.exc_info()[0]))
     finally:
         del mxd

     return dsList


def list_layers_to_xls(mxdPath, mxdFolder, xlsFolder):
    print('\nThe mxd file: %s' % mxdPath)
    fname = os.path.basename(mxdPath)
    fdir = os.path.dirname(mxdPath)
    dsList = scan_layers_in_mxd(mxdPath)
    wbFolder = fdir.replace(mxdFolder, xlsFolder)
    if not os.path.exists(wbFolder):
        os.makedirs(wbFolder)
    wbFilePath = os.path.join(wbFolder, fname + ".xlsx")
    print('\nThe xlsx file: %s' % wbFilePath)
    write_to_workbook(wbFilePath, dsList)


def scan_mxd_in_folder(mxdFolder, xlsFolder, date_filters):
    # parse date filters
    lower_date = None
    upper_date = None
    if date_filters is not None:
        dfs = re.split('<', date_filters)
        if len(dfs) == 2:
            if dfs[0] is not None and len(dfs[0]) > 0:
                lower_date = datetime.datetime.strptime(dfs[0], "%Y/%m/%d")
            if dfs[1] is not None and len(dfs[1]) > 0:
                upper_date = datetime.datetime.strptime(dfs[1], "%Y/%m/%d")

    # walk through all files
    for root, dirs, files in os.walk(mxdFolder):
        for fname in files:
            if fname.endswith(".mxd"):
                mxdPath = os.path.join(root, fname)
                modified_time = datetime.datetime.fromtimestamp(os.stat(mxdPath).st_mtime)
                # check against the filter
                is_filter_met = True
                if lower_date is not None:
                    is_filter_met = modified_time >= lower_date
                if upper_date is not None:
                    is_filter_met = is_filter_met and modified_time <= upper_date
                # work on the mxd file
                if is_filter_met == True:
                    list_layers_to_xls(mxdPath, mxdFolder, xlsFolder)


def scan_missed_mxds(mxdFolder, xlsFolder):
    missedMxds = []
    # walk through all files
    for root, dirs, files in os.walk(mxdFolder):
        for fname in files:
            if fname.endswith(".mxd"):
                mxdPath = os.path.join(root, fname)
                xlsPath = os.path.join(root.replace(mxdFolder, xlsFolder), fname + ".xlsx")
                if not os.path.exists(xlsPath):
                    mxdModifiedTime = datetime.datetime.fromtimestamp(os.stat(mxdPath).st_mtime)
                    missedMxd = {
                        "Path": mxdPath,
                        "ModTime": mxdModifiedTime.strftime('"%Y-%m-%d"')
                    }
                    missedMxds.append(missedMxd)
    missedMxds = sorted(missedMxds, key=lambda missedMxd: missedMxd['ModTime'])
    for missedMxd in missedMxds:
        print 'Found: %-10s  %s' % (missedMxd['ModTime'], missedMxd['Path'])
        list_layers_to_xls(missedMxd['Path'], mxdFolder, xlsFolder)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Scan mxds and list layers into spreadsheets')
    parser.add_argument('-m','--mxd', help='MXD Folder (input)', required=True)
    parser.add_argument('-x','--xls', help='XLS Folder (output)', required=True)
    parser.add_argument('-a','--action', help='Action Options (scan, comp, single)', required=False, default='scan')
    parser.add_argument('-f','--filter', help='Filter by dates. Ex. 2015/2/19<2016/2/18', required=False, default=None)
    parser.add_argument('-s','--single', help='A single MXD file (input for action: single', required=False, default=None)
    parser.add_argument('-c','--cfg', help='Config File', required=False, default=r'H:\MXD_Scan\config.xml')

    params = parser.parse_args()

    if params.cfg is not None:
        load_config(params.cfg)

    if params.action == 'scan':
        scan_mxd_in_folder(params.mxd, params.xls, params.filter)
    elif params.action == 'comp':
        scan_missed_mxds(params.mxd, params.xls)
    elif params.action == 'single':
        list_layers_to_xls(params.single, params.mxd, params.xls)
    else:
        print 'Error: unknown action [%s] for scanning' % params.action
