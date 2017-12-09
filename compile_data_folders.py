import sys
import os
import argparse
import tempfile
import shutil
#import xml.etree.ElementTree as ET
from lxml import etree as ET
import arcpy
from openpyxl import Workbook
from openpyxl import load_workbook
import logging

logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

XLS_TAB_NAME = "dataSource"
XLS_HEADERS = []

NETWORK_DRIVES = []

DATA_SOURCES = {}

XREF_TAB_NAME = "xRef"
LL_OUTPUT_HEADERS = ["Source Type", "Data Source", "Livelink Node Id"]
SDE_OUTPUT_HEADERS = ["Source Type", "Data Source", "SDE Path"]


def get_alias_path(lDrvPath):
    for nd in NETWORK_DRIVES:
        if lDrvPath.find(nd["Path"]) == 0:
            return lDrvPath.replace(nd["Path"], nd["Drive"])
        if lDrvPath.find(nd["Drive"]) == 0:
            return lDrvPath.replace(nd["Drive"], nd["Path"])
    return lDrvPath


def load_config(configFile):
    tree = ET.parse(configFile)
    root = tree.getroot()

    for child in root.iter('xlsTab'):
        # only one tab
        XLS_TAB_NAME = child.attrib['name']

    for child in root.iter('xlsHeader'):
        XLS_HEADERS.append(child.attrib['name'])
    # print XLS_HEADERS

    for child in root.iter('networkDrive'):
        NETWORK_DRIVES.append({
            'Drive': child.attrib['drive'],
            'Path': child.attrib['path']
        })
    # print NETWORK_DRIVES

    for child in root.iter('source'):
        nm = child.attrib['name']
        if nm not in DATA_SOURCES.keys():
            DATA_SOURCES[nm] = []
        for sc in child:
            DATA_SOURCES[nm].append({
                'LDrv':sc.attrib['path'],
                'Mode':sc.tag
            })
            DATA_SOURCES[nm].append({
                'LDrv':get_alias_path(sc.attrib['path']),
                'Mode':sc.tag
            })
    # print DATA_SOURCES


def get_source_type(lDrvPath):
    for k in DATA_SOURCES.keys():
        excluded = False
        ds = DATA_SOURCES[k]
        for cvt in [s for s in ds if s["Mode"] == "exclude"]:
            if lDrvPath.find(cvt["LDrv"]) == 0:
                excluded = True
                break
        if not excluded:
            for cvt in [s for s in ds if s["Mode"] == "add"]:
                if lDrvPath.find(cvt["LDrv"]) == 0:
                    return k
    return None


def read_from_workbook(wbPath, sheetName=None):
    wb = load_workbook(filename = wbPath, read_only=True)
    ws = wb[XLS_TAB_NAME]

    dsList = []
    # skip the first 2 rows
    r = 3
    hdrs = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

    while ws["A"+str(r)].value is not None:
        dsRecord = {}
        for c in range(0, len(XLS_HEADERS)):
            dsRecord[XLS_HEADERS[c]] = ws[hdrs[c]+str(r)].value
        dsList.append(dsRecord)
        r = r + 1

    wb.close()

    del wb

    return dsList


def write_to_workbook(wbPath, dsList, headerList):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = XREF_TAB_NAME

    # headers
    for c in range(0, len(headerList)):
        ws1.cell(row=1, column=c+1, value=headerList[c])

    # content
    s = 1 # skip the first 1 row
    for r in range(0, len(dsList)):
        for c in dsList[r]:
            # TODO: add style to cell?
            h = headerList.index(c)
            if h > -1:
                ws1.cell(row=r+s+1, column=h+1, value=dsList[r][headerList[h]])
            else:
                print('Invalid header [%s] in xls [%s]' % (c, wbPath))

    wb.save(filename = wbPath)
    wb.close()

    del wb


def parse_data_folder(dataSource):
    parts = dataSource.split('\\')
    idx = -1
    for p in list(reversed(range(0, len(parts)))):
        if parts[p].find('.') > -1:
            idx = p
            break
    return '\\'.join(parts[:idx])


def compile_data_folders_For_Livelink(xlsFolder, xlsOutput):
    dataFolderSet = set()
    for root, dirs, files in os.walk(xlsFolder):
        # walk through all files
        for fname in files:
            if fname.endswith(".xlsx") and not fname.startswith('~$'):
                # read from a xls file
                xlsPath = os.path.join(root, fname)
                print('\nThe xlsx file: %s' % xlsPath)
                lyrList = read_from_workbook(xlsPath)
                for lyr in lyrList:
                    #if lyr["Loaded?"] in ["LOADED", 'EXIST']:
                    if bool(lyr["Verified?"]) == True:
                        if lyr["Layer Type"] == "FeatureLayer":
                            df = parse_data_folder(lyr["Data Source"])
                            print('%s\t[%s]' % (df, lyr["Data Source"]))
                            dataFolderSet.add(df)

    dataFolderList = []
    for f in dataFolderSet:
        srcType = get_source_type(f)
        if srcType is not None:
            df = {
                "Source Type": srcType,
                "Data Source": f
            }
            dataFolderList.append(df)

    write_to_workbook(xlsOutput, sorted(dataFolderList, key=lambda df:df["Source Type"]), LL_OUTPUT_HEADERS)


def compile_data_sources_For_SDE(xlsFolder, xlsOutput):
    None


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Compile all data folders or sources into one spreadsheet')
    parser.add_argument('-x','--xls', help='XLS Folder (input)', required=True)
    parser.add_argument('-o','--dsf', help='XLS File (output)', required=True)
    parser.add_argument('-t','--tgt', help="Target Options (Livelink, SDE)", required=True)
    parser.add_argument('-c','--cfg', help='Config File', required=False, default=r'H:\MXD_Scan\config.xml')

    params = parser.parse_args()

    if os.path.exists(params.dsf):
        print('**** No Overwrite the existing output XLS file: %s' % xlsOutput)
        exit(-1)
    if params.cfg is not None:
        load_config(params.cfg)

    if params.tgt == "Livelink":
        compile_data_folders_For_Livelink(params.xls, params.dsf)
        print('**** The compiled data folder list for Livelink: %s' % xlsOutput)
    elif params.tgt == "SDE":
        compile_data_sources_For_SDE(params.xls, params.dsf)
        print('**** The compiled data source list for SDE: %s' % xlsOutput)

